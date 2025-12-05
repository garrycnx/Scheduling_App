# roster.py
import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
from openpyxl import Workbook

# ----------------------------------------------------
# Helper: Assign 2 consecutive OFF days
# ----------------------------------------------------
def assign_off_days(num_agents, dates):
    roster = {}
    n_days = len(dates)

    for a in range(1, num_agents + 1):
        off_start = random.randint(0, n_days - 2)   # 2-day OFF
        off_days = {dates[off_start], dates[off_start + 1]}
        roster[f"Agent_{a}"] = {d: ("OFF" if d in off_days else None) for d in dates}
    return roster


# ----------------------------------------------------
# auto shift generator based on candidate start hours
# ----------------------------------------------------
def generate_shift(start_hour, shift_length):
    end_hour = (start_hour + shift_length) % 24
    return f"{start_hour:02d}:00-{end_hour:02d}:00"


# ----------------------------------------------------
# Build weekly roster AUTO
# ----------------------------------------------------
def build_weekly_roster_auto(
    weekly_intervals,
    agents=None,
    out_path="weekly_roster.xlsx",
    shift_length_hours=9,
    start_hours=list(range(24))
):
    # ----------------------------------------------------
    # Step 1: Extract all unique dates
    # ----------------------------------------------------
    date_set = set()
    for r in weekly_intervals:
        try:
            dt = r["interval_start"].split(" ")[0]
        except:
            continue
        date_set.add(dt)

    dates = sorted(list(date_set), key=lambda x: datetime.strptime(x, "%d-%m-%Y"))

    # ----------------------------------------------------
    # Step 2: Determine required agents = max intervals
    # ----------------------------------------------------
    interval_df = pd.DataFrame(weekly_intervals)
    interval_df["date"] = interval_df["interval_start"].apply(lambda x: x.split(" ")[0])

    req_agents = interval_df.groupby("date")["required_after_shrinkage"].max().max()
    req_agents = int(np.ceil(req_agents))

    # If agents not provided → generate required number
    if agents is None:
        num_agents = req_agents
        agents = [f"Agent_{i}" for i in range(1, num_agents + 1)]
    else:
        num_agents = len(agents)

    # ----------------------------------------------------
    # Step 3: Assign OFF days
    # ----------------------------------------------------
    roster = assign_off_days(num_agents, dates)

    # ----------------------------------------------------
    # Step 4: Assign shifts for working days
    # ----------------------------------------------------
    for agent in roster:
        for d in dates:
            if roster[agent][d] == "OFF":
                continue
            # choose random shift start
            sh = random.choice(start_hours)
            roster[agent][d] = generate_shift(sh, shift_length_hours)

    # ----------------------------------------------------
    # Step 5: Build roster DataFrame
    # ----------------------------------------------------
    rows = []
    for agent in roster:
        row = {"agent": agent}
        for d in dates:
            row[d] = roster[agent][d]
        rows.append(row)

    roster_df = pd.DataFrame(rows)

    # ----------------------------------------------------
    # Step 6: Build Shift Pool summary
    # ----------------------------------------------------
    shift_pool = {}
    for agent in roster:
        for d in dates:
            sh = roster[agent][d]
            if sh != "OFF":
                shift_pool[sh] = shift_pool.get(sh, 0) + 1
    sp_df = pd.DataFrame(list(shift_pool.items()), columns=["shift", "count"])

    # ----------------------------------------------------
    # Step 7: Build Inflex table
    # ----------------------------------------------------
    temp = interval_df.copy()
    temp["rostered"] = req_agents
    temp["inflex"] = temp["rostered"] - temp["required_after_shrinkage"]
    inflex_df = temp

    # ----------------------------------------------------
    # Step 8: Export Excel
    # ----------------------------------------------------
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Roster"

    # Write roster sheet
    ws1.append(["agent"] + dates)
    for i, row in roster_df.iterrows():
        ws1.append([row["agent"]] + [row[d] for d in dates])

    # Shift Pool sheet
    ws2 = wb.create_sheet("Shift_Pool")
    ws2.append(["shift", "count"])
    for _, r in sp_df.iterrows():
        ws2.append([r["shift"], int(r["count"])])

    # Inflex sheet
    ws3 = wb.create_sheet("Inflex")
    ws3.append(list(inflex_df.columns))
    for _, r in inflex_df.iterrows():
        ws3.append(list(r.values))

    wb.save(out_path)

    return out_path, roster_df, inflex_df, sp_df
